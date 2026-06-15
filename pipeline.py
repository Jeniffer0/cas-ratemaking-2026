"""
================================================================================
CAS E-FORUM 2026 CALL PAPER

Enhancing Insurance Ratemaking Under Data Constraints:
A Machine Learning Framework for Market-Level Rate Adequacy Benchmarking
Using Aggregate Regulatory Data from the Kenyan General Insurance Market

Author  : Jeniffer Nasike Atetwe
          BSc Actuarial Science | OTHM Level 7 Diploma in Risk Management | International Diploma in Insurance Management | (TASK)

License : Mozilla Public License 2.0 (MPL 2.0)
          https://www.mozilla.org/en-US/MPL/2.0/

Repository : https://github.com/Jeniffer0/cas-ratemaking-2026

--------------------------------------------------------------------------------
OVERVIEW
--------------------------------------------------------------------------------
This pipeline replicates the full methodology of the paper in nine steps:

  Step 1  Data ingestion from IRA Kenya Excel workbooks (2023 and 2024)
  Step 2  Long-format dataset construction (insurer x class x year)
  Step 3  Feature engineering (nine features from aggregate statistics)
  Step 4  Exploratory data analysis and charts
  Step 5  Temporal holdout model training and evaluation
  Step 6  Full-dataset LightGBM for SHAP and signal generation
  Step 7  Buhlmann-Straub credibility-ML hybrid
  Step 8  SHAP analysis and visualisation
  Step 9  Decision framework output and summary statistics

All outputs (figures, CSV tables, JSON summary) are saved to /outputs/.

--------------------------------------------------------------------------------
INSTALLATION
--------------------------------------------------------------------------------
  pip install pandas numpy scikit-learn lightgbm shap matplotlib seaborn openpyxl

--------------------------------------------------------------------------------
DATA SETUP
--------------------------------------------------------------------------------
Place the two IRA Kenya Excel workbooks in a /data/ folder:

  data/IRA_Kenya_Annual_Statistics_2023.xlsx
  data/IRA_Kenya_Annual_Statistics_2024.xlsx

Update FILE_2023, FILE_2024, and APPENDIX_MAP below if your file names or
sheet names differ from the defaults.

--------------------------------------------------------------------------------
REPLICATION FOR OTHER JURISDICTIONS
--------------------------------------------------------------------------------
This pipeline is jurisdiction-agnostic. To apply it to NAIC Annual Statement
aggregates or any equivalent regulatory publication:

  1. Update FILE_2023 / FILE_2024 to point to your data files.
  2. Update APPENDIX_MAP with the sheet names in your publication.
  3. Update CLASSES with the line-of-business names used in your jurisdiction
     (see Table 1 of the paper for the IRA Kenya to ISO/NAIC class mapping).
  4. Run the pipeline unchanged.

================================================================================
"""

# ── Imports ───────────────────────────────────────────────────────────────────
import os
import json
import warnings

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import seaborn as sns
import openpyxl
import lightgbm as lgb
import shap

from sklearn.linear_model import ElasticNetCV
from sklearn.ensemble import RandomForestRegressor
from sklearn.preprocessing import LabelEncoder
from sklearn.metrics import mean_absolute_error, r2_score

warnings.filterwarnings("ignore")
os.makedirs("outputs", exist_ok=True)

plt.rcParams.update({
    "font.family": "Arial", "font.size": 11,
    "axes.titlesize": 12, "axes.titleweight": "bold",
    "figure.dpi": 150,
    "axes.spines.top": False, "axes.spines.right": False,
})

# ==============================================================================
# CONFIGURATION  -- update these for your data files and jurisdiction
# ==============================================================================

FILE_2023 = "data/IRA_Kenya_Annual_Statistics_2023.xlsx"
FILE_2024 = "data/IRA_Kenya_Annual_Statistics_2024.xlsx"

# 14 IRA Kenya general insurance classes (see Table 1 of paper for ISO/NAIC mapping)
CLASSES = [
    "Aviation", "Engineering", "Fire Domestic", "Fire Industrial",
    "Liability", "Marine", "Motor Private", "Motor Commercial",
    "Motor Commercial PSV", "Personal Accident", "Theft",
    "Workmens Compensation", "Medical", "Miscellaneous",
]

# Sheet names by year and data type
# Update these if your publication uses different sheet names
APPENDIX_MAP = {
    2023: {
        "gross_premium":       "APPENDIX 19",
        "outward_reinsurance": "APPENDIX 22",
        "net_earned_premium":  "APPENDIX 38",
        "net_incurred_claims": "APPENDIX 40",
        "uw_profit":           "APPENDIX 42",
        "pl_account":          "APPENDIX 1",
        "balance_sheets":      ["APPENDIX 5 I", "APPENDIX 5 II",
                                 "APPENDIX 5 III", "APPENDIX5 IV"],
    },
    2024: {
        "gross_premium":       "APPENDIX 21",
        "outward_reinsurance": "APPENDIX 24",
        "net_earned_premium":  "APPENDIX 40",
        "net_incurred_claims": "APPENDIX 42",
        "uw_profit":           "APPENDIX 44",
        "pl_account":          "APPENDIX 1",
        "balance_sheets":      ["APPENDIX 6 I", "APPENDIX 6 II",
                                 "APPENDIX 6 III", "APPENDIX 6 IV"],
    },
}

# Entities to exclude from all modelling
EXCLUDE_KEYWORDS = [
    "CONTINENTAL REINSURANCE", "EAST AFRICA REINSURANCE",
    "EAST AFRICAN REINSURANCE", "GHANA REINSURANCE",
    "KENYA REINSURANCE", "WAICA REINSURANCE", "TRANSRE",
    "AMOUNTS IN THOUSAND", "INSURERS", "REINSURERS",
    "TOTAL", "GRAND TOTAL",
]

# Model hyperparameters (Section 5 of paper)
RF_PARAMS = dict(n_estimators=300, max_depth=5, min_samples_leaf=5,
                 random_state=42, n_jobs=-1)
LGBM_PARAMS = dict(n_estimators=200, learning_rate=0.05, max_depth=4,
                   min_child_samples=5, reg_alpha=0.1, reg_lambda=0.1,
                   random_state=42, verbose=-1)

# Winsorisation percentiles for ICR target variable (Section 3.3)
ICR_WIN_LO = 0.05
ICR_WIN_HI = 0.95

# Bootstrap iterations for LightGBM MAE confidence interval (Section 5.2)
N_BOOTSTRAP = 1000

# Feature names for modelling (Section 4)
FEATURES = [
    "uw_margin",
    "log_gross_premium",
    "yoy_growth",
    "premium_mix",
    "market_share_class",
    "outward_reinsurance_ratio",
    "investment_income_ratio",
    "capital_adequacy_proxy",
    "class_enc",
]

# Human-readable labels for SHAP plots
FEATURE_LABELS = [
    "UW Margin",
    "Log Gross Premium (Size)",
    "YoY Premium Growth",
    "Premium Mix",
    "Market Share (Class)",
    "Outward Reinsurance Ratio",
    "Investment Income Ratio",
    "Capital Adequacy Proxy",
    "Class of Business",
]

TARGET = "loss_ratio_w"

# Plot colours
BLUE, RED, ORANGE, GREY = "#2E75B6", "#C00000", "#E47226", "#D9D9D9"


# ==============================================================================
# STEP 1: DATA INGESTION
# ==============================================================================

def _clean_name(name):
    """Return cleaned insurer name, or None if the row should be excluded."""
    if not isinstance(name, str):
        return None
    name = name.strip()
    if len(name) < 3:
        return None
    name_upper = name.upper()
    for kw in EXCLUDE_KEYWORDS:
        if kw in name_upper:
            return None
    return name


def read_wide_appendix(workbook, sheet_name, n_classes=14):
    """
    Read a wide-format insurer-by-class appendix.

    Layout assumed:
      Rows 0-3  : headers / title rows (skipped)
      Row 4+    : one insurer per row
      Column 1  : insurer name
      Columns 2 to 2+n_classes : one numeric value per class (CLASSES order)

    Returns pd.DataFrame with columns ['insurer'] + CLASSES.
    """
    try:
        ws = workbook[sheet_name]
    except KeyError:
        print(f"  WARNING: sheet '{sheet_name}' not found.")
        return pd.DataFrame(columns=["insurer"] + CLASSES)

    data = []
    for i, row in enumerate(ws.iter_rows(max_row=100, values_only=True)):
        if i < 4:
            continue
        name = _clean_name(row[1])
        if name is None:
            continue
        vals = list(row[2: 2 + n_classes])
        vals_clean = [
            float(v) if isinstance(v, (int, float)) and not isinstance(v, bool)
            else np.nan
            for v in vals
        ]
        vals_clean += [np.nan] * (n_classes - len(vals_clean))
        data.append([name] + vals_clean[:n_classes])

    return pd.DataFrame(data, columns=["insurer"] + CLASSES)


def read_pl_account(workbook, sheet_name):
    """
    Read the insurer-level profit and loss account.
    Extracts investment income (column index 3).

    Returns pd.DataFrame with columns ['insurer', 'investment_income'].
    """
    try:
        ws = workbook[sheet_name]
    except KeyError:
        print(f"  WARNING: sheet '{sheet_name}' not found.")
        return pd.DataFrame(columns=["insurer", "investment_income"])

    data = []
    for i, row in enumerate(ws.iter_rows(max_row=100, values_only=True)):
        if i < 4:
            continue
        name = _clean_name(row[1])
        if name is None:
            continue
        v = row[3] if len(row) > 3 else None
        inv = float(v) if isinstance(v, (int, float)) and not isinstance(v, bool) else np.nan
        data.append({"insurer": name, "investment_income": inv})

    return pd.DataFrame(data)


def read_balance_sheets(workbook, sheet_names):
    """
    Read insurer-level balance sheets (may be split across multiple sheet parts).
    Balance sheets are transposed: companies are columns, metrics are rows.
    Extracts 'Total Equity' for each insurer.

    Returns pd.DataFrame with columns ['insurer', 'total_equity'].
    """
    all_records = {}

    for sheet_name in sheet_names:
        try:
            ws = workbook[sheet_name]
        except KeyError:
            continue

        insurer_headers = None
        equity_row = None

        for i, row in enumerate(ws.iter_rows(max_row=60, values_only=True)):
            if i == 3:
                insurer_headers = [str(v).strip() if v else "" for v in row[2:]]
            if i >= 4 and row[1] and "total equity" in str(row[1]).lower():
                if equity_row is None:
                    equity_row = list(row[2:])

        if insurer_headers is None or equity_row is None:
            continue

        for j, raw_name in enumerate(insurer_headers):
            name = _clean_name(raw_name)
            if name is None or name in all_records:
                continue
            v = equity_row[j] if j < len(equity_row) else None
            eq = float(v) if isinstance(v, (int, float)) and not isinstance(v, bool) else np.nan
            all_records[name] = {"insurer": name, "total_equity": eq}

    return pd.DataFrame(list(all_records.values()))


def load_year(file_path, year):
    """Load all required appendices for a single year from one Excel workbook."""
    print(f"  Loading {year} from: {file_path}")
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    m = APPENDIX_MAP[year]
    return {
        "gross_premium":       read_wide_appendix(wb, m["gross_premium"]),
        "outward_reinsurance": read_wide_appendix(wb, m["outward_reinsurance"]),
        "net_earned_premium":  read_wide_appendix(wb, m["net_earned_premium"]),
        "net_incurred_claims": read_wide_appendix(wb, m["net_incurred_claims"]),
        "uw_profit":           read_wide_appendix(wb, m["uw_profit"]),
        "pl_account":          read_pl_account(wb, m["pl_account"]),
        "balance_sheets":      read_balance_sheets(wb, m["balance_sheets"]),
    }


# ==============================================================================
# STEP 2: DATASET CONSTRUCTION
# ==============================================================================

def build_long_dataset(data_dict, year):
    """
    Construct a long-format insurer-class-year dataset.

    Each row is one valid insurer-class-year observation.
    Validity requires: gross_premium > 0 AND net_earned_premium > 0.

    Parameters
    ----------
    data_dict : dict of DataFrames from load_year()
    year      : int

    Returns pd.DataFrame with raw financials for all valid observations.
    """
    prem   = data_dict["gross_premium"]
    outre  = data_dict["outward_reinsurance"]
    nep    = data_dict["net_earned_premium"]
    nic    = data_dict["net_incurred_claims"]
    uwp    = data_dict["uw_profit"]
    pl_df  = data_dict["pl_account"]
    bal_df = data_dict["balance_sheets"]

    market_total = {cls: prem[cls].sum(skipna=True) for cls in CLASSES}

    rows = []
    for _, prem_row in prem.iterrows():
        insurer = prem_row["insurer"]
        ins_total = sum(
            prem_row[c] for c in CLASSES
            if not pd.isna(prem_row[c]) and prem_row[c] > 0
        )

        pl_match  = pl_df[pl_df["insurer"] == insurer]
        bal_match = bal_df[bal_df["insurer"] == insurer]
        inv = pl_match["investment_income"].values[0]  if len(pl_match)  > 0 else np.nan
        eq  = bal_match["total_equity"].values[0]      if len(bal_match) > 0 else np.nan

        for cls in CLASSES:
            gp = prem_row[cls]
            if pd.isna(gp) or gp <= 0:
                continue

            def _get(df, col):
                m = df[df["insurer"] == insurer]
                return m[col].values[0] if len(m) > 0 else np.nan

            ne = _get(nep,   cls)
            ni = _get(nic,   cls)
            uw = _get(uwp,   cls)
            ot = _get(outre, cls)

            if pd.isna(ne) or ne <= 0:
                continue

            rows.append({
                "insurer":                   insurer,
                "class":                     cls,
                "year":                      year,
                "gross_premium":             gp,
                "net_earned_premium":        ne,
                "net_incurred_claims":       ni,
                "uw_profit":                 uw,
                "outward_reinsurance_prem":  ot,
                "investment_income":         inv,
                "total_equity":              eq,
                "loss_ratio":                ni / ne * 100 if not pd.isna(ni) else np.nan,
            })

    return pd.DataFrame(rows)


# ==============================================================================
# STEP 3: FEATURE ENGINEERING
# ==============================================================================

def engineer_features(df):
    """
    Construct the nine model features from the long-format dataset.

    Features (Section 4 of paper):
      1. uw_margin                  Underwriting margin
      2. log_gross_premium          Log gross premium (insurer size proxy)
      3. yoy_growth                 Year-on-year premium growth
      4. premium_mix                Class share of insurer total premium
      5. market_share_class         Insurer share of class market premium
      6. outward_reinsurance_ratio  Outward reinsurance as fraction of gross premium
      7. investment_income_ratio    Investment income as fraction of insurer total premium
      8. capital_adequacy_proxy     Total equity / net earned premium
      9. class_enc                  Label-encoded class of business

    Also constructs:
      loss_ratio_w    Winsorised ICR (modelling target)
      loss_ratio_lag  2023 ICR as lag for 2024 (constructed but not used in
                      primary model due to missingness for 2023 training obs;
                      available for future work with 3+ years of data)

    Returns (df_with_features, icr_lo, icr_hi)
    """
    df = df.copy()

    # Target variable: winsorise ICR
    lo = df["loss_ratio"].quantile(ICR_WIN_LO)
    hi = df["loss_ratio"].quantile(ICR_WIN_HI)
    df["loss_ratio_w"] = df["loss_ratio"].clip(lower=lo, upper=hi)

    # Feature 1: Underwriting margin
    df["uw_margin"] = df["uw_profit"] / df["gross_premium"]

    # Feature 2: Log gross premium
    df["log_gross_premium"] = np.log1p(df["gross_premium"])

    # Feature 3: Year-on-year premium growth (requires both years in df)
    pivot = df.pivot_table(
        index=["insurer", "class"], columns="year", values="gross_premium"
    ).reset_index()
    if 2023 in pivot.columns and 2024 in pivot.columns:
        pivot["yoy_growth"] = (pivot[2024] - pivot[2023]) / pivot[2023]
        pivot["yoy_growth"] = (
            pivot["yoy_growth"]
            .replace([np.inf, -np.inf], np.nan)
            .clip(-1.0, 2.0)
        )
        df = df.merge(pivot[["insurer", "class", "yoy_growth"]],
                      on=["insurer", "class"], how="left")
    else:
        df["yoy_growth"] = np.nan

    # Feature 4: Premium mix
    ins_total = df.groupby(["insurer", "year"])["gross_premium"].transform("sum")
    df["premium_mix"] = df["gross_premium"] / ins_total

    # Feature 5: Market share within class
    cls_total = df.groupby(["class", "year"])["gross_premium"].transform("sum")
    df["market_share_class"] = df["gross_premium"] / cls_total

    # Feature 6: Outward reinsurance ratio
    df["outward_reinsurance_ratio"] = (
        df["outward_reinsurance_prem"] / df["gross_premium"]
    )

    # Feature 7: Investment income ratio
    df["investment_income_ratio"] = df["investment_income"] / ins_total

    # Feature 8: Capital adequacy proxy
    df["capital_adequacy_proxy"] = df["total_equity"] / df["net_earned_premium"]

    # Feature 9: Class of business (label encoded)
    le = LabelEncoder()
    df["class_enc"] = le.fit_transform(df["class"])

    # Lagged loss ratio: constructed but NOT used in primary model
    # (undefined for all 2023 training observations; available for future work)
    lag = (
        df[df["year"] == 2023][["insurer", "class", "loss_ratio_w"]]
        .rename(columns={"loss_ratio_w": "loss_ratio_lag"})
    )
    df = df.merge(lag, on=["insurer", "class"], how="left")
    df.loc[df["year"] == 2023, "loss_ratio_lag"] = np.nan

    return df, lo, hi


# ==============================================================================
# STEP 4: EDA CHARTS
# ==============================================================================

def plot_icr_by_class(df):
    """Figure 1: ICR distribution by class of business."""
    fig, ax = plt.subplots(figsize=(14, 6))
    order = (df.groupby("class")["loss_ratio_w"]
               .median().sort_values(ascending=False).index)
    sns.boxplot(data=df, x="class", y="loss_ratio_w", order=order,
                palette="Blues_r", linewidth=0.8, ax=ax, fliersize=2)
    ax.axhline(100, color=RED, linewidth=1.2, linestyle="--", label="100% break-even")
    ax.set_xticklabels(ax.get_xticklabels(), rotation=40, ha="right", fontsize=9)
    ax.set_xlabel("Class of Business")
    ax.set_ylabel("Winsorised ICR (%)")
    ax.set_title("Figure 1. Distribution of Winsorised ICR by Class\n"
                 "Kenyan General Insurance Market, 2023-2024")
    ax.legend(fontsize=9)
    plt.tight_layout()
    plt.savefig("outputs/fig1_icr_by_class.png", dpi=300)
    plt.close()
    print("  Saved: outputs/fig1_icr_by_class.png")


def plot_icr_heatmap(df, year):
    """Figure 2/3: Insurer-by-class ICR heatmap for a single year."""
    pivot = df[df["year"] == year].pivot_table(
        index="insurer", columns="class", values="loss_ratio_w"
    )
    fig, ax = plt.subplots(figsize=(18, 10))
    sns.heatmap(pivot, cmap="RdYlGn_r", center=100, vmin=0, vmax=160,
                linewidths=0.3, linecolor="white",
                cbar_kws={"label": "Winsorised ICR (%)"},
                ax=ax, annot=False)
    ax.set_title(f"Figure. Insurer-by-Class ICR Heatmap ({year})\n"
                 "Red = high loss ratio | Green = low loss ratio")
    ax.set_xlabel("Class of Business")
    ax.set_ylabel("Insurer")
    ax.set_xticklabels(ax.get_xticklabels(), rotation=40, ha="right", fontsize=8)
    ax.set_yticklabels(ax.get_yticklabels(), fontsize=7)
    plt.tight_layout()
    plt.savefig(f"outputs/fig_heatmap_{year}.png", dpi=300)
    plt.close()
    print(f"  Saved: outputs/fig_heatmap_{year}.png")


# ==============================================================================
# STEP 5: MODEL TRAINING AND TEMPORAL HOLDOUT VALIDATION
# ==============================================================================

def train_and_evaluate(df_model):
    """
    Train Elastic Net, Random Forest, LightGBM under temporal holdout.
    Train on 2023 (N=308); evaluate on 2024 (N=308).

    Returns (results_dict, predictions_dict, fitted_models_dict).
    """
    train = df_model["year"] == 2023
    test  = df_model["year"] == 2024

    X_tr = df_model.loc[train, FEATURES].values
    y_tr = df_model.loc[train, TARGET].values
    X_te = df_model.loc[test,  FEATURES].values
    y_te = df_model.loc[test,  TARGET].values

    print(f"  Train (2023): {train.sum()} | Test (2024): {test.sum()}")

    # Elastic Net
    enet = ElasticNetCV(cv=5, max_iter=10000, random_state=42)
    enet.fit(X_tr, y_tr)
    yp_e = enet.predict(X_te)

    # Random Forest
    rf = RandomForestRegressor(**RF_PARAMS)
    rf.fit(X_tr, y_tr)
    yp_r = rf.predict(X_te)

    # LightGBM
    lgbm = lgb.LGBMRegressor(**LGBM_PARAMS)
    lgbm.fit(X_tr, y_tr)
    yp_l = lgbm.predict(X_te)

    # Bootstrap 95% CI for LightGBM MAE
    np.random.seed(42)
    boot_maes = []
    for _ in range(N_BOOTSTRAP):
        idx = np.random.choice(len(y_te), len(y_te), replace=True)
        boot_maes.append(mean_absolute_error(y_te[idx], yp_l[idx]))
    ci_lo, ci_hi = np.percentile(boot_maes, [2.5, 97.5])

    results = {
        "n_train": int(train.sum()),
        "n_test":  int(test.sum()),
        "enet":  {"mae": round(mean_absolute_error(y_te, yp_e), 2),
                  "r2":  round(r2_score(y_te, yp_e), 4)},
        "rf":    {"mae": round(mean_absolute_error(y_te, yp_r), 2),
                  "r2":  round(r2_score(y_te, yp_r), 4)},
        "lgbm":  {"mae": round(mean_absolute_error(y_te, yp_l), 2),
                  "r2":  round(r2_score(y_te, yp_l), 4),
                  "ci":  [round(ci_lo, 2), round(ci_hi, 2)]},
    }
    predictions = {"y_test": y_te, "enet": yp_e, "rf": yp_r, "lgbm": yp_l}
    models = {"enet": enet, "rf": rf, "lgbm": lgbm}
    return results, predictions, models


def plot_actual_vs_predicted(predictions):
    """Figure 4: Actual vs predicted ICR, all three models."""
    fig, axes = plt.subplots(1, 3, figsize=(15, 5))
    y_test = predictions["y_test"]
    for ax, (name, key) in zip(axes, [
        ("Elastic Net (Baseline)", "enet"),
        ("Random Forest", "rf"),
        ("LightGBM", "lgbm")
    ]):
        yp  = predictions[key]
        mae = mean_absolute_error(y_test, yp)
        r2  = r2_score(y_test, yp)
        ax.scatter(y_test, yp, alpha=0.5, s=25, color=BLUE,
                   edgecolors="white", linewidths=0.3)
        lims = [min(y_test.min(), yp.min()) - 5, max(y_test.max(), yp.max()) + 5]
        ax.plot(lims, lims, color=RED, linewidth=1.2, linestyle="--")
        ax.set_xlabel("Actual ICR (%)")
        ax.set_ylabel("Predicted ICR (%)")
        ax.set_title(f"{name}\nMAE={mae:.1f}pp | R\u00b2={r2:.3f}")
        ax.set_xlim(lims); ax.set_ylim(lims)
    plt.suptitle("Figure 4. Actual vs Predicted ICR — 2024 Holdout",
                 fontweight="bold", y=1.01)
    plt.tight_layout()
    plt.savefig("outputs/fig4_actual_vs_predicted.png", dpi=300, bbox_inches="tight")
    plt.close()
    print("  Saved: outputs/fig4_actual_vs_predicted.png")




# ==============================================================================
# STEP 5b: ROBUSTNESS ANALYSIS — EXCLUDING UNDERWRITING MARGIN
# ==============================================================================

def run_robustness(df_model):
    """
    Robustness check: estimate all models excluding underwriting margin.
    Documents the extent of predictor-outcome dependence (Section 6.2 of paper).

    Returns dict with performance metrics for both specifications.
    """
    FEATURES_NO_UW = [f for f in FEATURES if f != 'uw_margin']
    TARGET_COL = 'loss_ratio_w'

    train = df_model['year'] == 2023
    test  = df_model['year'] == 2024

    X_tr_no = df_model.loc[train, FEATURES_NO_UW].values
    y_tr    = df_model.loc[train, TARGET_COL].values
    X_te_no = df_model.loc[test,  FEATURES_NO_UW].values
    y_te    = df_model.loc[test,  TARGET_COL].values

    rf_no = RandomForestRegressor(**RF_PARAMS)
    rf_no.fit(X_tr_no, y_tr)
    yp_r_no = rf_no.predict(X_te_no)

    lgbm_no = lgb.LGBMRegressor(**LGBM_PARAMS)
    lgbm_no.fit(X_tr_no, y_tr)
    yp_l_no = lgbm_no.predict(X_te_no)

    return {
        'rf_no_uw':   {'mae': round(mean_absolute_error(y_te, yp_r_no), 2),
                       'r2':  round(r2_score(y_te, yp_r_no), 4)},
        'lgbm_no_uw': {'mae': round(mean_absolute_error(y_te, yp_l_no), 2),
                       'r2':  round(r2_score(y_te, yp_l_no), 4)},
    }

# ==============================================================================
# STEP 6: FULL-DATASET LIGHTGBM (for SHAP and signals)
# ==============================================================================

def fit_lgbm_full(df_model):
    """Fit LightGBM on the full dataset (2023 + 2024) for SHAP and signal generation."""
    X = df_model[FEATURES].values
    y = df_model[TARGET].values
    model = lgb.LGBMRegressor(**LGBM_PARAMS)
    model.fit(X, y)
    return model, X, y


# ==============================================================================
# STEP 7: BUHLMANN-STRAUB CREDIBILITY-ML HYBRID
# ==============================================================================

def compute_credibility_hybrid(df_model, lgbm_full, X_all):
    """
    Compute Buhlmann-Straub credibility-weighted hybrid predictions.

    Hybrid = Z * LightGBM_prediction + (1 - Z) * class_market_average_ICR

    Z(i,c) = v(i,c) / (v(i,c) + k)

    where v(i,c) = gross premium (credibility volume measure)
          k       = within-variance / between-variance (method of moments)

    Returns df_model with columns: pred_lgbm, class_avg_icr, k_param, Z,
                                   pred_hybrid, signal, adequacy_flag.
    """
    df_model = df_model.copy()
    df_model["pred_lgbm"] = lgbm_full.predict(X_all)

    class_avg = df_model.groupby("class")[TARGET].mean()
    df_model["class_avg_icr"] = df_model["class"].map(class_avg)

    within_var  = ((df_model[TARGET] - df_model["pred_lgbm"]) ** 2).mean()
    between_var = df_model.groupby("class")[TARGET].var().mean()
    k_param = within_var / between_var if between_var > 0 else 1.0

    df_model["k_param"] = k_param
    df_model["Z"] = df_model["gross_premium"] / (df_model["gross_premium"] + k_param)
    df_model["pred_hybrid"] = (
        df_model["Z"] * df_model["pred_lgbm"]
        + (1 - df_model["Z"]) * df_model["class_avg_icr"]
    )
    df_model["signal"] = df_model[TARGET] - df_model["pred_hybrid"]

    def _flag(s):
        if s > 20:  return "Potentially Underpriced"
        if s > 10:  return "Moderate Upward Pressure"
        if s > -5:  return "Within Market Range"
        if s > -10: return "Moderate Downward Pressure"
        return "Potentially Overpriced"

    df_model["adequacy_flag"] = df_model["signal"].apply(_flag)
    return df_model, k_param


def plot_credibility_weights(df_model):
    """Figure 5: Distribution of Buhlmann-Straub credibility weights."""
    fig, ax = plt.subplots(figsize=(10, 5))
    ax.hist(df_model["Z"], bins=40, color=BLUE, edgecolor="white", alpha=0.85)
    ax.axvline(0.5, color=RED, linestyle="--", linewidth=1.2, label="Z = 0.5")
    ax.set_xlabel("Credibility Weight Z")
    ax.set_ylabel("Number of Insurer-Class-Year Observations")
    ax.set_title("Figure 5. Distribution of Buhlmann-Straub Credibility Weights\n"
                 "Z = 1.0: full weight on ML prediction | Z = 0.0: full weight on class average")
    ax.legend()
    plt.tight_layout()
    plt.savefig("outputs/fig5_credibility_weights.png", dpi=300)
    plt.close()
    print("  Saved: outputs/fig5_credibility_weights.png")


def plot_rate_adequacy_signals(df_model):
    """Figure 6: Average rate adequacy signal by class."""
    class_signals = (df_model.groupby("class")["signal"]
                              .mean().sort_values(ascending=False))
    colors = [
        RED if v > 10 else ORANGE if v > 0 else BLUE if v < -10 else GREY
        for v in class_signals.values
    ]
    fig, ax = plt.subplots(figsize=(14, 6))
    ax.bar(class_signals.index, class_signals.values, color=colors, edgecolor="white")
    ax.axhline(0,   color="black", linewidth=0.8)
    ax.axhline(10,  color=RED,  linewidth=0.8, linestyle="--", alpha=0.5)
    ax.axhline(-10, color=BLUE, linewidth=0.8, linestyle="--", alpha=0.5)
    ax.set_xticklabels(class_signals.index, rotation=40, ha="right", fontsize=9)
    ax.set_ylabel("Average Rate Adequacy Signal (pp)")
    ax.set_title("Figure 6. Average Rate Adequacy Signal by Class\n"
                 "Signal = Actual ICR minus Hybrid Prediction "
                 "(positive = potential underpricing)")
    ax.legend(handles=[
        mpatches.Patch(color=RED,    label="Potential underpricing (> +10pp)"),
        mpatches.Patch(color=BLUE,   label="Potential overpricing (< -10pp)"),
        mpatches.Patch(color=GREY,   label="Within market range"),
    ], fontsize=9)
    plt.tight_layout()
    plt.savefig("outputs/fig6_rate_adequacy_signals.png", dpi=300)
    plt.close()
    print("  Saved: outputs/fig6_rate_adequacy_signals.png")


# ==============================================================================
# STEP 8: SHAP ANALYSIS
# ==============================================================================

def compute_and_plot_shap(lgbm_full, X_all):
    """Compute SHAP values and generate beeswarm and bar charts."""
    explainer   = shap.TreeExplainer(lgbm_full)
    shap_values = explainer.shap_values(X_all)

    mean_abs_shap = pd.Series(
        np.abs(shap_values).mean(axis=0), index=FEATURE_LABELS
    ).sort_values(ascending=False)

    # Bar chart (Figure 7)
    fig, ax = plt.subplots(figsize=(10, 6))
    colors = [BLUE if i == 0 else GREY for i in range(len(mean_abs_shap))]
    ax.barh(mean_abs_shap.index[::-1], mean_abs_shap.values[::-1],
            color=colors[::-1], edgecolor="white")
    ax.set_xlabel("Mean |SHAP Value| (percentage points)")
    ax.set_title("Figure 7. SHAP Feature Importance\n"
                 "Mean Absolute Impact on Predicted ICR (LightGBM, Full Dataset)")
    for i, val in enumerate(mean_abs_shap.values[::-1]):
        ax.text(val + 0.2, i, f"{val:.1f}", va="center", fontsize=9)
    plt.tight_layout()
    plt.savefig("outputs/fig7_shap_importance.png", dpi=300)
    plt.close()
    print("  Saved: outputs/fig7_shap_importance.png")

    # Beeswarm (Figure 8)
    plt.figure(figsize=(10, 7))
    shap.summary_plot(shap_values, X_all, feature_names=FEATURE_LABELS,
                      show=False, plot_type="dot")
    plt.title("Figure 8. SHAP Beeswarm Plot\n"
              "Distribution of Feature Contributions to Predicted ICR")
    plt.tight_layout()
    plt.savefig("outputs/fig8_shap_beeswarm.png", dpi=300, bbox_inches="tight")
    plt.close()
    print("  Saved: outputs/fig8_shap_beeswarm.png")

    return mean_abs_shap


# ==============================================================================
# STEP 9: OUTPUT TABLES AND SUMMARY
# ==============================================================================

def save_all_outputs(results, df_model, mean_abs_shap,
                     icr_lo, icr_hi, wlr_2023, wlr_2024):
    """Save all CSV tables and JSON summary statistics."""

    # Table 4: Model performance
    pd.DataFrame([
        {"Model": "Elastic Net (Baseline)",
         "MAE (pp)": results["enet"]["mae"], "R-Squared": results["enet"]["r2"],
         "Notes": "Linear baseline"},
        {"Model": "Random Forest",
         "MAE (pp)": results["rf"]["mae"],   "R-Squared": results["rf"]["r2"],
         "Notes": "Best out-of-sample performance"},
        {"Model": "LightGBM",
         "MAE (pp)": results["lgbm"]["mae"], "R-Squared": results["lgbm"]["r2"],
         "Notes": f"95% CI: [{results['lgbm']['ci'][0]}, {results['lgbm']['ci'][1]}]pp"},
    ]).to_csv("outputs/table4_model_performance.csv", index=False)

    # Table 5: Class-level rate adequacy signals
    cls_sig = (df_model.groupby("class")["signal"]
                        .agg(avg_signal="mean", max_signal="max",
                             min_signal="min", n="count")
                        .round(2)
                        .sort_values("avg_signal", ascending=False)
                        .reset_index())
    cls_sig.to_csv("outputs/table5_class_signals.csv", index=False)

    # Table 6: SHAP importance
    shap_df = mean_abs_shap.reset_index()
    shap_df.columns = ["Feature", "Mean |SHAP| (pp)"]
    shap_df["Rank"] = range(1, len(shap_df) + 1)
    shap_df[["Rank", "Feature", "Mean |SHAP| (pp)"]].to_csv(
        "outputs/table6_shap_importance.csv", index=False)

    # Table 7: Decision framework
    pd.DataFrame([
        {"> +20", "Large writer", "Both years",
         "Investigate pricing; benchmark against class average; consider corrective action"},
        {"> +20", "Small writer", "Either year",
         "Apply credibility hybrid; seek supplementary policy-level data"},
        {"+10 to +20", "Any", "Both years",
         "Review claims development; flag for next actuarial review cycle"},
        {"+10 to +20", "Any", "One year only",
         "Monitor; may reflect idiosyncratic large claim"},
        {"-5 to +10", "Any", "Any",
         "Within expected market variation; no action required"},
        {"< -10", "Large writer", "Both years",
         "Investigate; determine if outperformance reflects pricing or underwriting"},
        {"< -10", "Small writer", "Either year",
         "Directional signal only; corroborate with other data sources"},
    ]).to_csv("outputs/table7_decision_framework.csv", index=False)

    # Full insurer-class signal output
    df_model[[
        "insurer", "class", "year", "loss_ratio_w", "pred_hybrid",
        "signal", "adequacy_flag", "Z", "gross_premium"
    ]].sort_values("signal", ascending=False).to_csv(
        "outputs/full_rate_adequacy_signals.csv", index=False)

    # JSON summary (all key stats for paper)
    summary = {
        "data": {
            "n_obs":           len(df_model),
            "n_train":         results["n_train"],
            "n_test":          results["n_test"],
            "icr_win_lo":      round(icr_lo, 1),
            "icr_win_hi":      round(icr_hi, 1),
            "wlr_2023":        round(wlr_2023, 1),
            "wlr_2024":        round(wlr_2024, 1),
        },
        "model_performance": results,
        "credibility": {
            "k_param": round(df_model["k_param"].iloc[0], 4),
            "mean_Z":  round(df_model["Z"].mean(), 4),
        },
        "shap": {f: round(v, 3) for f, v in mean_abs_shap.items()},
        "signal_range": {
            "min": round(df_model["signal"].min(), 1),
            "max": round(df_model["signal"].max(), 1),
        },
    }
    with open("outputs/pipeline_summary.json", "w") as f:
        json.dump(summary, f, indent=2)

    print("  Saved: outputs/table4_model_performance.csv")
    print("  Saved: outputs/table5_class_signals.csv")
    print("  Saved: outputs/table6_shap_importance.csv")
    print("  Saved: outputs/table7_decision_framework.csv")
    print("  Saved: outputs/full_rate_adequacy_signals.csv")
    print("  Saved: outputs/pipeline_summary.json")


# ==============================================================================
# MAIN
# ==============================================================================

def main():
    print("=" * 70)
    print("CAS RATEMAKING CALL PAPER 2026")
    print("IRA Kenya Market-Level Rate Adequacy Benchmarking Pipeline")
    print("=" * 70)

    # Step 1: Ingest data
    print("\n[Step 1] Loading data...")
    data_2023 = load_year(FILE_2023, 2023)
    data_2024 = load_year(FILE_2024, 2024)

    # Step 2: Build dataset
    print("\n[Step 2] Building dataset...")
    df23 = build_long_dataset(data_2023, 2023)
    df24 = build_long_dataset(data_2024, 2024)
    df_raw = pd.concat([df23, df24], ignore_index=True)
    print(f"  Raw: {len(df_raw)} obs | 2023: {len(df23)} | 2024: {len(df24)}")
    print(f"  Insurers: {df23['insurer'].nunique()} in 2023, "
          f"{df24['insurer'].nunique()} in 2024")

    wlr_2023 = df23["net_incurred_claims"].sum() / df23["net_earned_premium"].sum() * 100
    wlr_2024 = df24["net_incurred_claims"].sum() / df24["net_earned_premium"].sum() * 100
    print(f"  Weighted market ICR: 2023 = {wlr_2023:.1f}%, 2024 = {wlr_2024:.1f}%")

    # Step 3: Feature engineering
    print("\n[Step 3] Engineering features...")
    df, icr_lo, icr_hi = engineer_features(df_raw)
    df_model = df.dropna(subset=FEATURES + [TARGET]).copy()
    print(f"  Winsorised ICR: [{icr_lo:.1f}%, {icr_hi:.1f}%]")
    print(f"  Modelling dataset: {len(df_model)} observations")

    # Step 4: EDA
    print("\n[Step 4] EDA charts...")
    plot_icr_by_class(df_model)
    plot_icr_heatmap(df_model, 2023)
    plot_icr_heatmap(df_model, 2024)

    # Step 5: Train and evaluate
    print("\n[Step 5] Temporal holdout validation...")
    results, predictions, fitted_models = train_and_evaluate(df_model)
    print(f"\n  Elastic Net:   MAE={results['enet']['mae']:.1f}pp  "
          f"R2={results['enet']['r2']:.3f}")
    print(f"  Random Forest: MAE={results['rf']['mae']:.1f}pp  "
          f"R2={results['rf']['r2']:.3f}")
    print(f"  LightGBM:      MAE={results['lgbm']['mae']:.1f}pp  "
          f"R2={results['lgbm']['r2']:.3f}  "
          f"95%CI=[{results['lgbm']['ci'][0]:.1f},{results['lgbm']['ci'][1]:.1f}]pp")
    plot_actual_vs_predicted(predictions)

    # Step 6: Full-dataset LightGBM
    print("\n[Step 6] Full-dataset LightGBM...")
    lgbm_full, X_all, y_all = fit_lgbm_full(df_model)

    # Step 7: Credibility hybrid
    print("\n[Step 7] Credibility-ML hybrid...")
    df_model, k_param = compute_credibility_hybrid(df_model, lgbm_full, X_all)
    print(f"  Buhlmann k = {k_param:.4f} | Mean Z = {df_model['Z'].mean():.4f}")
    print(f"  Signal range: [{df_model['signal'].min():.1f}, "
          f"{df_model['signal'].max():.1f}]pp")
    plot_credibility_weights(df_model)
    plot_rate_adequacy_signals(df_model)

    # Step 8: SHAP
    print("\n[Step 8] SHAP analysis...")
    mean_abs_shap = compute_and_plot_shap(lgbm_full, X_all)
    print(f"\n  Feature importance (ranked):")
    for feat, val in mean_abs_shap.items():
        print(f"    {feat:35s}: {val:.3f}pp")

    # Step 9: Save all outputs
    print("\n[Step 9] Saving outputs...")
    save_all_outputs(results, df_model, mean_abs_shap,
                     icr_lo, icr_hi, wlr_2023, wlr_2024)

    print("\n" + "=" * 70)
    print("PIPELINE COMPLETE — all outputs in /outputs/")
    print("=" * 70)
    print("""
Figures:
  fig1_icr_by_class.png            ICR distribution by class (EDA)
  fig_heatmap_2023.png             Insurer-class heatmap (2023)
  fig_heatmap_2024.png             Insurer-class heatmap (2024)
  fig4_actual_vs_predicted.png     Model fit: all three models
  fig5_credibility_weights.png     Buhlmann-Straub Z distribution
  fig6_rate_adequacy_signals.png   Rate adequacy signals by class
  fig7_shap_importance.png         SHAP bar chart
  fig8_shap_beeswarm.png           SHAP beeswarm plot

Tables:
  table4_model_performance.csv     MAE / R-squared comparison
  table5_class_signals.csv         Rate adequacy signals by class
  table6_shap_importance.csv       SHAP feature importance ranking
  table7_decision_framework.csv    Decision framework
  full_rate_adequacy_signals.csv   Per-insurer-class signals

Summary:
  pipeline_summary.json            All key statistics for paper
    """)


if __name__ == "__main__":
    main()
